from flask import Flask, render_template, request
import qrcode
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

def device_used(ip, session):
    wb = load_workbook(attendance_file)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == session and row[4] == ip:
            return True
    return False

def already_marked(roll_no, session):
    wb = load_workbook(attendance_file)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == session and row[1] == roll_no:
            return True
    return False


def save_attendance(session, roll, name, ip):
    wb = load_workbook(attendance_file)
    ws = wb.active
    time = datetime.now().strftime("%H:%M:%S")
    ws.append([session, roll, name, time, ip])
    wb.save(attendance_file)

app = Flask(__name__)

attendance_file = "attendance.xlsx"
current_session = None

def init_excel():
    if not os.path.exists(attendance_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Session", "Roll No", "Name", "Time", "Device IP"])
        wb.save(attendance_file)

init_excel()
    
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/teacher')
def teacher():
    global current_session
    now = datetime.now().strftime("%Y-%m-%d_%H-%M")
    current_session = "CLASS_" + now

    qr_link = "http://192.168.0.160:5000/student"
    img = qrcode.make(qr_link)

    if not os.path.exists("static"):
        os.mkdir("static")

    img.save("static/qr.png")

    return render_template('teacher.html', session=current_session)


@app.route('/records')
def records():
    wb = load_workbook(attendance_file)
    ws = wb.active
    data = list(ws.values)
    return render_template('records.html', records=data)

@app.route('/student', methods=['GET', 'POST'])
def student():
    global current_session

    if request.method == 'POST':
        roll = request.form['roll']
        name = request.form['name']
        ip = request.remote_addr

        if device_used(ip, current_session):
            message = "This device already marked attendance!"
        elif already_marked(roll, current_session):
            message = "Attendance already marked for this class!"
        else:
            save_attendance(current_session, roll, name, ip)
            message = "Attendance marked successfully!"

        return render_template('success.html', message=message)

    return render_template('student.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)